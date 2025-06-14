# -*- coding: utf-8 -*--
import json
import os
import re
import shutil
from datetime import datetime

import openpyxl
import pyramid.httpexceptions as exc
from infolica.exceptions.custom_error import CustomError
from infolica.models.constant import Constant
from infolica.models.models import (Affaire, AffaireNumero, Cadastre, Facture,
                                    Numero, NumeroDiffere, NumeroEtat,
                                    NumeroEtatHisto, NumeroType, NumeroRelation,
                                    VNumeros, VNumerosAffaires)
from infolica.scripts.authentication import check_connected
from infolica.scripts.utils import Utils
from pyramid.view import view_config
from sqlalchemy import BigInteger, Integer, and_, func
from sqlalchemy.dialects.postgresql import ARRAY, aggregate_order_by


def _update_numero_etat(request, numero_id, etat_id):
    num = request.dbsession.query(Numero).get(numero_id)
    num.etat_id = etat_id

    num_histo = NumeroEtatHisto()
    num_histo.numero_id = numero_id
    num_histo.numero_etat_id = etat_id
    num_histo.date = datetime.today()
    return num



@view_config(route_name='numeros', request_method='GET', renderer='json')
@view_config(route_name='numeros_s', request_method='GET', renderer='json')
def numeros_view(request):
    """
    Return all numeros
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()

    numero = int(request.params['numero']) if 'numero' in request.params else None
    cadastre_id = [int(a) for a in request.params['cadastre_id'].split(",")] if 'cadastre_id' in request.params else None
    type_id = [int(a) for a in request.params['type_id'].split(",")] if 'type_id' in request.params else None
    etat_id = [int(a) for a in request.params['etat_id'].split(",")] if 'etat_id' in request.params else None

    query = request.dbsession.query(VNumeros)

    if numero:
        query = query.filter(VNumeros.numero == numero)
    if cadastre_id:
        query = query.filter(VNumeros.cadastre_id.in_(cadastre_id))
    if type_id:
        query = query.filter(VNumeros.type_numero_id.in_(type_id))
    if etat_id:
        query = query.filter(VNumeros.etat_id.in_(etat_id))

    query = query.order_by(VNumeros.numero.asc()).all()

    return Utils.serialize_many(query)


@view_config(route_name='types_numeros', request_method='GET', renderer='json')
@view_config(route_name='types_numeros_s', request_method='GET', renderer='json')
def types_numeros_view(request):
    """
    Return all types_numeros
    """
    immeule_thr = request.registry.settings['numero_type_immeuble_thr']
    query = request.dbsession.query(NumeroType).filter(NumeroType.ordre < immeule_thr).order_by(NumeroType.ordre.asc()).all()
    return Utils.serialize_many(query)


@view_config(route_name='types_numeros_mo', request_method='GET', renderer='json')
def types_numeros_mo_view(request):
    """
    Return all types_numeros
    """
    immeule_thr = request.registry.settings['numero_type_immeuble_thr']
    query = request.dbsession.query(NumeroType).filter(NumeroType.ordre >= immeule_thr).order_by(NumeroType.ordre.asc()).all()
    return Utils.serialize_many(query)


@view_config(route_name='etats_numeros', request_method='GET', renderer='json')
@view_config(route_name='etats_numeros_s', request_method='GET', renderer='json')
def etats_numeros_view(request):
    """
    Return all etats_numeros
    """
    id_artefact = request.registry.settings['numero_artefact_id']
    query = request.dbsession.query(NumeroEtat).filter(NumeroEtat.id != id_artefact).all()
    return Utils.serialize_many(query)


@view_config(route_name='numero_by_id', request_method='GET', renderer='json')
def numeros_by_id_view(request):
    """
    Return numeros by id
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()

    # Get controle mutation id
    id = request.id = request.matchdict['id']
    query = request.dbsession.query(VNumeros).filter(
        VNumeros.id == id).first()
    return Utils.serialize_one(query)


@view_config(route_name='recherche_numeros', request_method='POST', renderer='json')
@view_config(route_name='recherche_numeros_s', request_method='POST', renderer='json')
def numeros_search_view(request):
    """
    Search numeros
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()

    settings = request.registry.settings
    search_limit = int(settings['search_limit'])

    params = request.params

    matDiff = False
    if 'matDiff' in params:
        matDiff = True if params['matDiff'] == 'true' else False

    # Set conditions
    conditions = Utils.get_search_conditions(VNumeros, params, ignore_params=["matDiff"])

    # filter by conditions
    query = request.dbsession.query(VNumeros).order_by(
        VNumeros.cadastre,
        VNumeros.numero.desc()
    ).filter(*conditions)

    query = query.filter(VNumeros.type_numero_id <= 4)

    # if option matDiff selected
    if matDiff:
        query = query.filter(
            and_(
                VNumeros.diff_entree.isnot(None),
                VNumeros.diff_sortie.is_(None)
                )
            )

    query = query.limit(search_limit).all()
    return Utils.serialize_many(query)


@view_config(route_name='numeros', request_method='POST', renderer='json')
@view_config(route_name='numeros_s', request_method='POST', renderer='json')
def numeros_new_view(request, params=None):
    """
    Add new numeros
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    if not params:
        params = request.params
        params['no_access'] = "_".join(str(params['cadastre_id']), str(params['numero']))

    # tester si le numéro à réserver est plus grand que le max +1 déjà existant dans la base. Si oui, répondre par une erreur
    num_max = request.dbsession.query(func.max(Numero.numero)).filter(Numero.cadastre_id == params['cadastre_id']).scalar()
    if int(params['numero']) > num_max + 1:
        raise CustomError(CustomError.NUMBER_REGISTRATION_FAILED.format(params['numero'], params['cadastre_id'], num_max))

    # nouveau numero
    record = Numero()
    record = Utils.set_model_record(record, params)

    request.dbsession.add(record)
    request.dbsession.flush()

    Utils.get_data_save_response(
            Constant.SUCCESS_SAVE.format(Numero.__tablename__)
    )

    return record.id


@view_config(route_name='numeros', request_method='PUT', renderer='json')
@view_config(route_name='numeros_s', request_method='PUT', renderer='json')
def numeros_update_view(request):
    """
    Update numeros
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    # Get numero id
    id = request.params['id'] if 'id' in request.params else None

    # Get numero record
    record = request.dbsession.query(Numero).filter(
        Numero.id == id).first()

    if not record:
        raise CustomError(
            CustomError.RECORD_WITH_ID_NOT_FOUND.format(Numero.__tablename__, id))

    # update numero state in numero_history if changed
    last_record_etat_id = record.etat_id
    record = Utils.set_model_record(record, request.params)

    if 'etat_id' in request.params:
        if request.params['etat_id'] != last_record_etat_id:
            params = Utils._params(
                numero_id=record.id, numero_etat_id=request.params['etat_id'])
            numeros_etat_histo_new_view(request, params)

    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(Numero.__tablename__))


@view_config(route_name='delete_numero', request_method='DELETE', renderer='json')
def numeros_delete_view(request):
    """
    Delete numeros
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    # Get numero id
    # id = request.params['id'] if 'id' in request.params else None
    numero_id = request.matchdict['id']
    affaire_id = request.params.get("affaire_id")

    # get affaire to check if it is already sent
    aff = request.dbsession.query(Affaire).get(affaire_id)
    if aff.date_envoi is not None:
        raise CustomError(CustomError.SENT_AFFAIRE_EXCEPTION.format(affaire_id))

    # make sure that bf to delete is reserved in current affaire
    check_reserved_bf = request.dbsession.query(AffaireNumero).filter(
        AffaireNumero.numero_id==numero_id,
        AffaireNumero.affaire_id==affaire_id,
        AffaireNumero.type_id==request.registry.settings['affaire_numero_type_nouveau_id'],
    ).first()
    if check_reserved_bf is None:
        raise CustomError(f"Impossible de supprimer le numéro {numero_id}, il ne figure pas dans la liste des biens-fonds réservés pour l'affaire {affaire_id}")

    cadastre = request.dbsession.query(Cadastre).join(Numero).filter(Numero.id==numero_id).first()
    cadastre = cadastre.nom if cadastre is not None else None

    # get numero_relation and remove it
    nr = request.dbsession.query(NumeroRelation).filter(
        NumeroRelation.numero_id_associe==numero_id,
        NumeroRelation.affaire_id==affaire_id
    ).all()
    for tmp in nr:
        request.dbsession.delete(tmp)

    an = request.dbsession.query(AffaireNumero).filter(
        AffaireNumero.numero_id==numero_id,
    )
    nb_affaires = an.count()

    affnum = an.filter(AffaireNumero.affaire_id==affaire_id).all()
    for tmp in affnum:
        request.dbsession.delete(tmp)

    num = request.dbsession.query(Numero).filter(
            Numero.id==numero_id
    ).first()

    if nb_affaires == 1:
        # get numero_etat_histo and remove it
        neh = request.dbsession.query(NumeroEtatHisto).filter(
            NumeroEtatHisto.numero_id==numero_id
        ).all()
        for tmp in neh:
            request.dbsession.delete(tmp)

        # get numero_differe and remove it
        nd = request.dbsession.query(NumeroDiffere).filter(
            NumeroDiffere.numero_id==numero_id,
        ).all()
        for tmp in nd:
            request.dbsession.delete(tmp)

        # remove numero
        request.dbsession.delete(num)

        message = f"Le numéro {num.numero} (cadastre: {cadastre}) a été correctement délié de l'affaire et supprimé, car il n'était lié à aucune autre affaire."

    else:
        an_ = an.all()
        message = f"Le numéro {num.numero} (cadastre: {cadastre}) a été correctement délié de l'affaire.<br>Il reste lié aux affaires suivantes: {', '.join([str(an.affaire_id) for an in an_])}."


    return {"message": message}


@view_config(route_name='numero_by_id', request_method='DELETE', renderer='json')
def numeros_by_id_delete_view(request):
    """
    Supprimer/abandonner numeros by id
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    settings = request.registry.settings
    projet_id = int(settings['numero_projet_id'])
    abandonne_id = int(settings['numero_abandonne_id'])

    # Get numero by id
    id = request.matchdict['id']
    query = request.dbsession.query(Numero).filter(
        Numero.id == id).first()

    if query:
        if query.etat_id == projet_id:
            query.etat_id = abandonne_id
        elif query.etat_id == abandonne_id:
            query.etat_id = projet_id

    return Utils.get_data_save_response(Constant.SUCCESS_DELETE.format(Numero.__tablename__))

###########################################################
# NUMERO ETAT HISTO
###########################################################


@view_config(route_name='numeros_etat_histo', request_method='POST', renderer='json')
@view_config(route_name='numeros_etat_histo_s', request_method='POST', renderer='json')
def numeros_etat_histo_new_view(request, params=None):
    """
    Add new numero_etat_histo
    """
    # Check authorization
    settings = request.registry.settings
    if not Utils.has_permission(request, settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    if not params:
        params = request.params

    # check if etat already exists
    numEtatHisto = request.dbsession.query(NumeroEtatHisto).filter(
        NumeroEtatHisto.numero_id == params['numero_id']
    ).filter(
        NumeroEtatHisto.numero_etat_id == params['numero_etat_id']
    ).first()

    if numEtatHisto is not None:
        return

    # nouveau numero
    record = NumeroEtatHisto()
    record = Utils.set_model_record(record, params)

    request.dbsession.add(record)
    request.dbsession.flush()

    Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(
        NumeroEtatHisto.__tablename__)
    )

    return record.id

###########################################################
# AFFAIRE-NUMERO
###########################################################


@view_config(route_name='affaire_numeros_by_affaire_id', request_method='GET', renderer='json')
def affaire_numeros_view(request):
    """
    numero_affaire
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()
    conditions = list()
    affaire_id = request.matchdict["id"]

    # Récupérer les id des numéros de la MO
    settings = request.registry.settings

    numeros_immeubles_type_id = [
        settings['numero_bf_id'],
        settings['numero_ddp_id'],
        settings['numero_ppe_id'],
        settings['numero_pcop_id']
    ]

    conditions.append(VNumerosAffaires.affaire_id == affaire_id)
    conditions.append(VNumerosAffaires.numero_type_id.in_(numeros_immeubles_type_id))

    if 'affaire_numero_actif' in request.params:
        affaire_numero_actif = request.params['affaire_numero_actif']
        affaire_numero_actif = True if affaire_numero_actif.upper() == 'TRUE' else False
        conditions.append(VNumerosAffaires.affaire_numero_actif == affaire_numero_actif)

    records = request.dbsession.query(VNumerosAffaires).filter(
        *conditions).order_by(VNumerosAffaires.numero.desc()).all()
    return Utils.serialize_many(records)


@view_config(route_name='affaire_numeros', request_method='PUT', renderer='json')
@view_config(route_name='affaire_numeros_s', request_method='PUT', renderer='json')
def affaire_numero_update_view(request):
    """
    Update numeros_affaires
    """
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    affnum_id = request.params["id"] if "id" in request.params else None

    record = request.dbsession.query(AffaireNumero).filter(AffaireNumero.id == affnum_id).first()

    if not record:
        raise CustomError(
            CustomError.RECORD_WITH_ID_NOT_FOUND.format(AffaireNumero.__tablename__, affnum_id))

    record = Utils.set_model_record(record, request.params)

    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(AffaireNumero.__tablename__))


@view_config(route_name='affaire_numeros', request_method='DELETE', renderer='json')
@view_config(route_name='affaire_numeros_s', request_method='DELETE', renderer='json')
def affaire_numero_delete_view(request):
    """
    delete numeros_affaires
    """
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    affnum_id = None
    affaire_id = None
    numero_id = None
    record = None
    if "id" in request.params:
        affnum_id = request.params["id"]
        record = request.dbsession.query(AffaireNumero).filter(AffaireNumero.id == affnum_id).first()
    elif "affaire_id" in request.params and "numero_id" in request.params:
        affaire_id = int(request.params["affaire_id"])
        numero_id = int(request.params["numero_id"])
        record = request.dbsession.query(AffaireNumero).filter(and_(
            AffaireNumero.affaire_id == affaire_id,
            AffaireNumero.numero_id == numero_id
        )).first()
    else:
        raise CustomError(CustomError.INCOMPLETE_REQUEST)

    if not record:
        raise CustomError(
            CustomError.RECORD_WITH_ID_NOT_FOUND.format(AffaireNumero.__tablename__, affnum_id))


    # Get affaire_type_id and config values
    affaire_type_id = request.dbsession.query(Affaire).filter(Affaire.id == affaire_id).first().type_id
    affaire_type_cadastration_id = int(request.registry.settings['affaire_type_cadastration_id'])

    # supprimer la facture du numéro si c'est une affaire de cadastration
    if affaire_type_id == affaire_type_cadastration_id:
        factNum = request.dbsession.query(Facture).filter(and_(
            Facture.affaire_id == affaire_id,
            Facture.numeros.op("@>")("{" + str(numero_id) + "}")
        )).all()

        for fact in factNum:
            request.dbsession.delete(fact)



    request.dbsession.delete(record)

    return Utils.get_data_save_response(Constant.SUCCESS_DELETE.format(AffaireNumero.__tablename__))


@view_config(route_name='affaire_numeros', request_method='POST', renderer='json')
@view_config(route_name='affaire_numeros_s', request_method='POST', renderer='json')
def affaire_numero_new_view(request, params=None):
    """
    Add new affaire-numero
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    if not params:
        params = request.params

    # check if affaire_numero relation already exists
    affaireNumero = request.dbsession.query(AffaireNumero).filter(
        AffaireNumero.affaire_id == params['affaire_id']
    ).filter(
        AffaireNumero.numero_id == params['numero_id']
    ).filter(
        AffaireNumero.type_id == params['type_id']
    ).first()

    if affaireNumero is not None:
        return

    # nouveau affaire_numero
    record = AffaireNumero()
    record = Utils.set_model_record(record, params)

    request.dbsession.add(record)

    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(AffaireNumero.__tablename__))


@view_config(route_name='add_affaire_numero', request_method='POST', renderer='json')
def add_affaire_numero_view(request):
    """
    Add new affaire-numeros
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    numeros = json.loads(request.params["numeros"]) if "numeros" in request.params else None
    affaire_id = request.params["affaire_id"] if "affaire_id" in request.params else None

    numeros.sort(reverse=False, key=lambda x: int(x.split("_")[1]))

    for num in numeros:
        cadastre_id, numero_bf = [int(a) for a in num.split("_")]

        # check if numero already exists in DB
        numero = request.dbsession.query(Numero).filter(Numero.cadastre_id==cadastre_id, Numero.numero==numero_bf).first()
        # get actual max number of cadastre
        actual_max_number = request.dbsession.query(func.max(Numero.numero)).filter(Numero.cadastre_id==cadastre_id).scalar()

        if numero is None and numero_bf <= (actual_max_number + 1):
            # create numero
            numero_bf_id = request.registry.settings["numero_bf_id"]
            numero_projet_id = request.registry.settings["numero_projet_id"]
            params = {
                "cadastre_id": cadastre_id,
                "numero": numero_bf,
                "type_id": numero_bf_id,
                "etat_id": numero_projet_id,
            }
            numero = Utils.addNewRecord(request, Numero, params)

            # create numero_etat
            params = {
                "numero_id": numero.id,
                "numero_etat_id": numero_projet_id,
                "date": datetime.now().date(),
            }
            Utils.addNewRecord(request, NumeroEtatHisto, params)

        elif numero_bf > (actual_max_number + 1):
            raise f"Le numero de BF candidat ({numero_bf}) dépasse le prochain numéro disponible ({actual_max_number + 1}) pour le cadastre n° {cadastre_id}"


        # check if affaire-numero already exists in DB
        affaire_numero_type_nouveau_id = request.registry.settings["affaire_numero_type_nouveau_id"]
        affaire_numero = request.dbsession.query(AffaireNumero).filter(AffaireNumero.numero_id==numero.id, AffaireNumero.affaire_id==affaire_id, AffaireNumero.type_id==affaire_numero_type_nouveau_id).first()
        if affaire_numero is None:
            params = {
                "affaire_id": affaire_id,
                "numero_id": numero.id,
                "type_id": affaire_numero_type_nouveau_id,
                "actif": True,
            }
            affaire_numero = Utils.addNewRecord(request, AffaireNumero, params)

    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(AffaireNumero.__tablename__))


###########################################################
# NUMERO- AFFAIRE
###########################################################

@view_config(route_name='numero_affaires_by_numero_id', request_method='GET', renderer='json')
def numeros_affaire_view(request):
    """
    Get affaires by numero_id
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()

    numero_id = request.matchdict['id']

    query = request.dbsession.query(VNumerosAffaires).filter(
        VNumerosAffaires.numero_id == numero_id).all()

    return Utils.serialize_many(query)


###########################################################
# NUMERO DIFFERE
###########################################################

@view_config(route_name='numeros_differes', request_method='GET', renderer='json')
def numero_differe_view(request):
    """
    get numero_differe
    """
    # Check connected
    if not check_connected(request):
        raise exc.HTTPForbidden()

    numero_projet_id = int(request.registry.settings['numero_projet_id'])
    numero_vigueur_id = int(request.registry.settings['numero_vigueur_id'])

    numero_etat_vigueur_id = request.registry.settings['numero_vigueur_id']

    role = request.params['role'] if 'role' in request.params else None

    num_agg = func.array_agg(VNumeros.numero, type_=ARRAY(Integer))
    num_id_agg = func.array_agg(VNumeros.id, type_=ARRAY(Integer))
    diff_id_agg = func.array_agg(VNumeros.diff_id, type_=ARRAY(Integer))
    numeros_vigueur_check = func.bool_and(VNumeros.etat_id == numero_etat_vigueur_id)
    query = request.dbsession.query(
        VNumeros.diff_affaire_id,
        VNumeros.cadastre,
        num_agg,
        num_id_agg,
        diff_id_agg,
        func.min(VNumeros.diff_entree),
        VNumeros.diff_operateur_id,
        VNumeros.diff_operateur_nom,
        VNumeros.diff_operateur_prenom,
        VNumeros.diff_operateur_initiales,
        VNumeros.diff_req_ref,
        numeros_vigueur_check
    )

    if role == "mo":
        user_id = request.params['user_id'] if 'user_id' in request.params else None

        if user_id is not None:
            query = query.filter(VNumeros.diff_operateur_id == user_id)

        query = query.filter(and_(
            VNumeros.diff_entree.isnot(None),
            VNumeros.diff_sortie == None
        ))

    elif role == "secr":
        query = query.filter(and_(
            VNumeros.diff_req_radiation.isnot(True),
            VNumeros.diff_sortie.isnot(None),
            VNumeros.etat_id.in_((numero_projet_id, numero_vigueur_id))
        ))

    elif role == "coord":
        user_id = request.params['user_id'] if 'user_id' in request.params else None

        if user_id is not None:
            query = query.filter(VNumeros.diff_operateur_id == user_id)

        query = query.filter(and_(
            VNumeros.diff_sortie.isnot(None),
            VNumeros.diff_controle == None
        ))

    result = query.group_by(
        VNumeros.diff_affaire_id,
        VNumeros.cadastre,
        VNumeros.diff_operateur_id,
        VNumeros.diff_operateur_nom,
        VNumeros.diff_operateur_prenom,
        VNumeros.diff_operateur_initiales,
        VNumeros.diff_req_ref
    ).having(func.array_length(num_agg, 1) > 0).all()

    affaire_ready = True if ('affaire_ready' in request.params and request.params['affaire_ready'] == "true") else False

    numeros = []
    for num in result:
        if affaire_ready is True:
            if num[11] is True:
                numeros.append({
                    'diff_affaire_id': num[0],
                    'cadastre': num[1],
                    'numero': num[2],
                    'numero_id': num[3],
                    'diff_id': num[4],
                    'diff_entree': datetime.strftime(num[5], '%Y-%m-%d'),
                    'diff_operateur_id': num[6],
                    'diff_operateur_nom': num[7],
                    'diff_operateur_prenom': num[8],
                    'diff_operateur_initiales': num[9],
                    'diff_req_ref': num[10],
                    'numeros_vigueur_check': num[11]
                })
        else:
            numeros.append({
                'diff_affaire_id': num[0],
                'cadastre': num[1],
                'numero': num[2],
                'numero_id': num[3],
                'diff_id': num[4],
                'diff_entree': datetime.strftime(num[5], '%Y-%m-%d'),
                'diff_operateur_id': num[6],
                'diff_operateur_nom': num[7],
                'diff_operateur_prenom': num[8],
                'diff_operateur_initiales': num[9],
                'diff_req_ref': num[10],
                'numeros_vigueur_check': num[11]
            })


    return numeros


@view_config(route_name='numeros_differes', request_method='POST', renderer='json')
@view_config(route_name='numeros_differes_s', request_method='POST', renderer='json')
def numero_differe_new_view(request):
    """
    Add new numero_differe
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    # nouveau numero_differe
    record = NumeroDiffere()
    record = Utils.set_model_record(record, request.params)

    request.dbsession.add(record)

    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(NumeroDiffere.__tablename__))


@view_config(route_name='numeros_differes', request_method='PUT', renderer='json')
@view_config(route_name='numeros_differes_s', request_method='PUT', renderer='json')
def numero_differe_update_view(request):
    """
    Update numero_differe
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    if "numero_diff_id" in request.params:
        numdiff_id = request.params["numero_diff_id"]
        record = request.dbsession.query(NumeroDiffere).filter(NumeroDiffere.id == numdiff_id).first()

    if "numero_id" in request.params:
        num_id = request.params["numero_id"]
        record = request.dbsession.query(NumeroDiffere).filter(NumeroDiffere.numero_id == num_id).first()

    # update numero_differe
    record = Utils.set_model_record(record, request.params)
    return Utils.get_data_save_response(Constant.SUCCESS_SAVE.format(NumeroDiffere.__tablename__))


@view_config(route_name='numeros_differes', request_method='DELETE', renderer='json')
def numero_differe_delete_view(request):
    """
    DELETE numero_differe
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    numero_id = request.params["numero_id"] if "numero_id" in request.params else None

    if numero_id is None:
        raise CustomError(CustomError.RECORD_WITH_ID_NOT_FOUND.format(Numero.__tablename__, numero_id))

    record = request.dbsession.query(NumeroDiffere).filter(NumeroDiffere.numero_id == numero_id).first()

    # delete numero_differe
    request.dbsession.delete(record)

    return Utils.get_data_save_response(Constant.SUCCESS_DELETE.format(NumeroDiffere.__tablename__))

#####################################
#  Numeros remaniement parcellaire  #
#####################################


def _getNumberId(request, numero, cadastre_id):
    numero = request.dbsession.query(
        Numero.id,
        Numero.type_id
    ).filter(
        Numero.cadastre_id == cadastre_id,
        Numero.numero == numero
    ).first()
    return (numero.id, numero.type_id) if numero is not None else (None, None)


def _getCadastre(request, cadastre_id):
    cadastre = request.dbsession.query(
        Cadastre.nom
    ).filter(
        Cadastre.id == cadastre_id
    ).first()
    return cadastre[0] if cadastre is not None else None


def _getAffairesIdFromNumeroId(request, numero_id, numero_type_id):
    affaires_id_agg = func.array_agg(AffaireNumero.affaire_id, type_=ARRAY(BigInteger))

    affaires_id = request.dbsession.query(
        affaires_id_agg
    ).filter(
        AffaireNumero.numero_id == numero_id,
        AffaireNumero.type_id == numero_type_id,
        AffaireNumero.actif == True
    ).group_by(AffaireNumero.numero_id).first()
    return [str(i) for i in affaires_id[0]] if affaires_id is not None else []


@view_config(route_name='loadfile_bf_rp', request_method='POST', renderer='json')
def loadfile_bf_rp(request):
    """
    Get File containing BF of RP and read it
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    affaire_id = request.params["affaire_id"] if "affaire_id" in request.params else None
    file = request.params["file"] if "file" in request.params else None

    if file is None:
        return exc.HTTPError('Le fichier est vide')

    temporary_directory = request.registry.settings['temporary_directory']
    affaire_numero_type_nouveau_id = request.registry.settings['affaire_numero_type_nouveau_id']
    numero_ddp_id = int(request.registry.settings['numero_ddp_id'])
    file_path = os.path.join(temporary_directory, file.filename)

    if os.path.exists(file_path):
        os.remove(file_path)

    with open(file_path, 'wb') as output_file:
        shutil.copyfileobj(file.file, output_file)

    wb = openpyxl.load_workbook(file_path)
    data = []

    numero_query = request.dbsession.query(Numero)

    # =============================
    # Let's focus on Infolica sheet
    # =============================
    sheet = 'Infolica'
    ws = wb[sheet]

    numero_id_agg = func.array_agg(Numero.id, type_=ARRAY(BigInteger))
    numero_numero_agg = func.array_agg(aggregate_order_by(Numero.numero, Numero.numero.asc()), type_=ARRAY(BigInteger))

    # prepare query to search Number
    data_query = request.dbsession.query(
        numero_id_agg,
        Numero.cadastre_id,
        numero_numero_agg,
        Cadastre.nom
    ).join(
        Cadastre, Cadastre.id == Numero.cadastre_id
    )

    data_id = []

    row_i = 2
    while row_i < 1000:
        numero_id = ws.cell(row=row_i, column=1).value

        if numero_id is not None:
            data_id.append(numero_id)
            row_i += 1

        else:
            results = data_query.filter(
                Numero.id.in_(data_id)
            ).group_by(Numero.cadastre_id, Cadastre.nom).order_by(Cadastre.nom).all()

            tmp = []
            for result in results: # parcourir les cadastres

                liste_numeros = []
                for i, numero_ in enumerate(result[2]):
                    num = numero_query.filter(Numero.id == result[0][i]).first()

                    error = False
                    font_color = 'black'
                    if num is not None and num.type_id == numero_ddp_id:
                        error = True
                        font_color = 'blue'

                    liste_numeros.append({
                        'numero_id': result[0][i],
                        'cadastre_id': result[1],
                        'numero': str(numero_) if error is False else 'DDP ' + str(numero_),
                        'cadastre': result[3],
                        'sheet': sheet,
                        'font_color': font_color,
                        'error': error,
                    })

                tmp.append({
                    'cadastre': result[3],
                    'cadastre_id': result[1],
                    'liste_numeros': liste_numeros
                })

            data.append({
                'source': sheet,
                'data': tmp
            })

            break

    # ===========================
    # Let's focus on Terris sheet
    # ===========================
    sheet = 'Terris'
    ws = wb[sheet]
    tmp = []

    row_i = 2
    while row_i < 1000:
        cadastre_id = ws.cell(row=row_i, column=2).value


        if cadastre_id is not None:
            numero = re.split('\D', str(ws.cell(row=row_i, column=3).value))[0]

            numero_id, numero_type_id = _getNumberId(request, numero, cadastre_id)
            affaires_id = []
            if numero_id is not None:
                affaires_id = _getAffairesIdFromNumeroId(request, numero_id, numero_type_id=affaire_numero_type_nouveau_id)

            cadastre = _getCadastre(request, cadastre_id)

            # errors
            error = False
            font_color = 'red'
            if affaire_id in (affaires_id):
                font_color = 'green'
            elif len(affaires_id) > 0:
                numero = numero + (" [affaire(s): " + ', '.join(affaires_id) + "]" if len(affaires_id) > 0 else "")
                font_color = 'blue'
                error = True

            if numero_type_id == numero_ddp_id:
                numero = 'DDP ' + numero
                font_color = 'blue'
                error = True


            numero_ = {
                'numero_id': numero_id,
                'cadastre_id': cadastre_id,
                'numero': numero,
                'cadastre': cadastre,
                'sheet': sheet,
                'font_color': font_color,
                'error': error
            }

            cadastre_id_already_exists = False
            for elem in tmp:
                if elem['cadastre_id'] == cadastre_id:
                    elem['liste_numeros'].append(numero_)
                    cadastre_id_already_exists = True
                    break

            if cadastre_id_already_exists is False:
                tmp.append({
                    'cadastre': cadastre,
                    'cadastre_id': cadastre_id,
                    'liste_numeros': [numero_]
                })

            row_i += 1

        else:

            for tmp_ln in tmp: # parcourir les cadastres
                tmp_ln['liste_numeros'].sort(key=lambda x: int(x['numero'].split(' ')[0]))

            data.append({
                'source': sheet,
                'data': tmp
            })

            break

    if os.path.exists(file_path):
        os.remove(file_path)

    return data


def __save_bf_rp(request, affaire_id, data, numero_type_id, numero_etat_id, numero_ancien_nouveau_id):
    num_req = request.dbsession.query(Numero)
    an_req = request.dbsession.query(AffaireNumero)

    date = datetime.strftime(datetime.now(), "%Y-%m-%d")

    for num_cad in data:
        # on parcourt les cadastres
        for numero_obj in num_cad['liste_numeros']:

            numero = str(numero_obj['numero']).split(' ')[0]

            num = num_req.filter(
                Numero.cadastre_id == num_cad['cadastre_id'],
                Numero.numero == numero
            ).first()

            if num is None:
                num = Numero(
                    cadastre_id = num_cad['cadastre_id'],
                    numero = numero,
                    etat_id = numero_etat_id,
                    type_id = numero_type_id
                )
                request.dbsession.add(num)
                request.dbsession.flush()

                # update numero_etat_histo
                neh = NumeroEtatHisto(
                    numero_id = num.id,
                    numero_etat_id = numero_etat_id,
                    date = date
                )

                request.dbsession.add(neh)

            else:
                # s'assurer que le numéro ait l'état en projet
                if not num.etat_id == numero_etat_id:
                    num.etat_id = numero_etat_id

                    # update numero_etat_histo
                    neh = NumeroEtatHisto(
                        numero_id = num.id,
                        numero_etat_id = numero_etat_id,
                        date = date
                    )

                    request.dbsession.add(neh)

            # log dans table affaire_numero if not already exists
            an = an_req.filter(
                AffaireNumero.affaire_id == affaire_id,
                AffaireNumero.numero_id == num.id,
                AffaireNumero.type_id == numero_ancien_nouveau_id
            ).first()

            if an is None:
                an = AffaireNumero(
                    affaire_id = affaire_id,
                    numero_id = num.id,
                    type_id = numero_ancien_nouveau_id,
                    actif = True
                )

                request.dbsession.add(an)
    return


@view_config(route_name='save_bf_rp', request_method='POST', renderer='json')
def save_bf_rp(request):
    """
    Save BF of RP-file
    """
    # Check authorization
    if not Utils.has_permission(request, request.registry.settings['affaire_numero_edition']):
        raise exc.HTTPForbidden()

    affaire_id = request.params["affaire_id"] if "affaire_id" in request.params else None
    num_projet = json.loads(request.params["num_projet"]) if "num_projet" in request.params else None
    num_vigueur = json.loads(request.params["num_vigueur"]) if "num_vigueur" in request.params else None

    affaire_numero_type_ancien_id = request.registry.settings['affaire_numero_type_ancien_id']
    affaire_numero_type_nouveau_id = request.registry.settings['affaire_numero_type_nouveau_id']
    numero_projet_id = request.registry.settings['numero_projet_id']
    numero_vigueur_id = request.registry.settings['numero_vigueur_id']
    numero_bf_id = request.registry.settings['numero_bf_id']

    __save_bf_rp(request, affaire_id, num_projet, numero_bf_id, numero_projet_id, affaire_numero_type_nouveau_id)
    __save_bf_rp(request, affaire_id, num_vigueur, numero_bf_id, numero_vigueur_id, affaire_numero_type_ancien_id)

    return exc.HTTPOk()
